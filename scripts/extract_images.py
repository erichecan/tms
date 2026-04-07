import openpyxl
import base64
import os
import psycopg2
import re
from io import BytesIO

DATABASE_URL = "postgresql://neondb_owner:npg_a0t9YKjwEkWP@ep-purple-block-ahewbbc2-pooler.c-3.us-east-1.aws.neon.tech/neondb?sslmode=require&channel_binding=require"

def get_base64_image(image):
    try:
        data = image._data()
        return f"data:image/png;base64,{base64.b64encode(data).decode('utf-8')}"
    except Exception as e:
        print(f"  [ERROR] Failed to convert image to base64: {e}")
        return None

def update_container_item_image(conn, container_no, fba_shipment_id, sku, image_base64):
    try:
        with conn.cursor() as cur:
            # Match by container_no and either fba_shipment_id or sku
            query = """
                UPDATE container_items ci
                SET details = jsonb_set(COALESCE(ci.details, '{}'), '{image}', %s)
                FROM containers c
                WHERE ci.container_id = c.id
                AND c.container_no = %s
                AND (
                    (ci.fba_shipment_id = %s AND %s != '') OR 
                    (ci.sku = %s AND %s != '')
                )
            """
            cur.execute(query, (f'"{image_base64}"', container_no, fba_shipment_id, fba_shipment_id, sku, sku))
            if cur.rowcount > 0:
                return True
    except Exception as e:
        print(f"  [ERROR] Database update failed: {e}")
    return False

def is_container_no(s):
    if not s: return False
    s = str(s).strip().upper()
    if len(s) < 4: return False
    # Common prefixes
    return any(s.startswith(p) for p in ["TCLU", "BMOU", "CBHU", "MSCU", "WHLU", "ONEU", "MAEU", "TRHU", "COSU", "EGP", "AMZ", "CCLU", "CSGU", "EGHU", "FFAU", "SMCU", "CAME"])

def process_excel(filepath):
    if not os.path.isfile(filepath):
        print(f"File not found: {filepath}")
        return

    conn = psycopg2.connect(DATABASE_URL)
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    total_updated = 0
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith("2024") or "已完成" in sheet_name:
            continue
            
        ws = wb[sheet_name]
        print(f"\nProcessing sheet: {sheet_name}")
        
        images_by_row = {}
        if hasattr(ws, '_images'):
            for img in ws._images:
                try:
                    row = img.anchor._from.row + 1
                    b64 = get_base64_image(img)
                    if b64 and row not in images_by_row:
                        images_by_row[row] = b64
                except Exception as e:
                    print(f"  [WARNING] Could not process image: {e}")
                    continue
        
        if not images_by_row:
            print("  No images found.")
            continue

        print(f"  Found {len(images_by_row)} images across different rows.")

        last_container_no = None
        for row_idx in range(1, ws.max_row + 1):
            c1 = ws.cell(row=row_idx, column=1).value
            c2 = ws.cell(row=row_idx, column=2).value
            if is_container_no(c1): last_container_no = str(c1).strip().upper()
            elif is_container_no(c2): last_container_no = str(c2).strip().upper()
            
            if row_idx in images_by_row:
                img_data = images_by_row[row_idx]
                if not last_container_no:
                    continue
                
                # Extract identifiers from the same row
                vals = [ws.cell(row=row_idx, column=c).value for c in range(1, 11)]
                fba = str(vals[4]).strip() if vals[4] else ""
                sku = str(vals[2]).strip() if vals[2] else ""
                
                if update_container_item_image(conn, last_container_no, fba, sku, img_data):
                    print(f"  [OK] Updated image for {last_container_no} / {fba or sku} at Row {row_idx}")
                    total_updated += 1
                else:
                    # Generic attempt if fba/sku not found but we have a container
                    update_container_item_image(conn, last_container_no, "", "", img_data)

    conn.commit()
    conn.close()
    wb.close()
    print(f"\nTotal items updated with images: {total_updated}")

if __name__ == "__main__":
    process_excel("2026年转运汇总表.xlsx")
