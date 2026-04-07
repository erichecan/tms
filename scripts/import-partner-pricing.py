#!/usr/bin/env python3
"""
Phase 2: 导入 合作单位报价卡.xlsx 结构化 sheet → partner_pricing_rules
用法:
  python scripts/import-partner-pricing.py --preview    # 预览，不写入
  python scripts/import-partner-pricing.py --commit     # 写入数据库
"""

import os
import re
import sys
import json
import argparse
from decimal import Decimal

import openpyxl
import psycopg2

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ENV_PATH = os.path.join(ROOT, "apps", "backend", ".env")
FILE_QUOTE = os.path.join(ROOT, "合作单位报价卡.xlsx")


def _load_database_url():
    url = os.environ.get("DATABASE_URL", "").strip()
    if not url or "..." in url:
        if os.path.isfile(ENV_PATH):
            with open(ENV_PATH, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if line.startswith("DATABASE_URL="):
                        val = line.split("=", 1)[1].strip().strip('"').strip("'")
                        url = val
                        break
    return url


def _clean(v):
    if v is None:
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return int(v) if v == int(v) else float(v)
    s = str(v).strip()
    return s if s else None


# ─── Tier definitions ─────────────────────────────────────────────────────────
# Map header text → (pallet_tier_min, pallet_tier_max, unit_type)
TIER_MAP = {
    "1-4板":   (1, 4, "per_trip"),
    "1-3板":   (1, 3, "per_trip"),
    "5-14板":  (5, 14, "per_trip"),
    "5-13板":  (5, 13, "per_trip"),
    "5-8板":   (5, 8, "per_trip"),
    "9-13板":  (9, 13, "per_trip"),
    "15-28板": (15, 28, "per_trip"),
    "14-28板": (14, 28, "per_trip"),
    "散板":    (1, 999, "per_pallet"),
    "散板/拼货": (1, 999, "per_pallet"),
    "整车":    (1, 28, "per_trip"),
}


def match_tier(header_text):
    """Match a header cell to a tier definition. Returns (label, min, max, unit_type) or None."""
    s = str(header_text).strip()
    # Direct match
    for label, (tmin, tmax, utype) in TIER_MAP.items():
        if label in s:
            return (label, tmin, tmax, utype)
    # Try extracting ranges like "5-14板(10P)"
    m = re.match(r"(\d+)-(\d+)", s)
    if m and "板" in s:
        return (s, int(m.group(1)), int(m.group(2)), "per_trip")
    if "散" in s or "拼" in s:
        return (s, 1, 999, "per_pallet")
    return None


def parse_price_cell(v):
    """Parse a price cell. Returns (base_price, unit_price, unit_type_override, notes).
    - Pure number: base_price=number
    - "10/板": unit_price=10, unit_type=per_pallet
    - "300（200起步）": base_price=300, notes="200起步"
    - Text with no number: notes only
    """
    if v is None:
        return None, None, None, None
    if isinstance(v, (int, float)):
        return float(v), None, None, None
    s = str(v).strip()
    if not s:
        return None, None, None, None

    # "10/板" or "25/P" pattern
    m = re.match(r"([\d.]+)\s*/\s*[板Pp点]", s)
    if m:
        return None, float(m.group(1)), "per_pallet", None

    # "300（200起步）" pattern
    m = re.match(r"([\d.]+)\s*[（(](.+?)[)）]", s)
    if m:
        return float(m.group(1)), None, None, m.group(2)

    # Pure number with commas
    cleaned = s.replace(",", "").replace(" ", "").replace("$", "")
    try:
        return float(cleaned), None, None, None
    except ValueError:
        # Text-only note (e.g., "拼货不单独发车，需等待顺路车")
        return None, None, None, s


def expand_destinations(dest_str):
    """Expand 'YYZ3/YYZ4/YYZ7' → ['YYZ3', 'YYZ4', 'YYZ7']."""
    if not dest_str:
        return []
    s = str(dest_str).strip().upper()
    # Find all warehouse codes (with or without trailing digit)
    codes = re.findall(r'[A-Z]{2,3}\d*[A-Z]?', s)
    # Filter out very short matches and common non-warehouse words
    codes = [c for c in codes if len(c) >= 3 and c not in ('AND', 'THE', 'FOR', 'NOT', 'ALL', 'UPS')]
    if codes:
        return codes
    # Try splitting by common separators
    parts = re.split(r'[/,，、\s]+', s)
    return [p.strip() for p in parts if p.strip() and len(p.strip()) >= 2]


# ─── Sheet Parsers ────────────────────────────────────────────────────────────

def parse_matrix_sheet(ws, sheet_name):
    """Parse standard matrix sheet (仓号 × pallet tiers). Returns list of rule dicts."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return []

    # Find header row with 仓号
    header_idx = -1
    for i in range(min(10, len(rows))):
        if rows[i]:
            for cell in rows[i]:
                if cell and "仓号" in str(cell):
                    header_idx = i
                    break
        if header_idx >= 0:
            break
    if header_idx < 0:
        return []

    header = rows[header_idx]
    # Find dest column
    dest_col = -1
    for i, v in enumerate(header or []):
        if v and "仓号" in str(v):
            dest_col = i
            break
    if dest_col < 0:
        return []

    # Find tier columns
    tier_cols = []  # [(col_idx, label, tier_min, tier_max, unit_type)]
    for i, v in enumerate(header or []):
        if v is None or i == dest_col:
            continue
        tier = match_tier(v)
        if tier:
            label, tmin, tmax, utype = tier
            tier_cols.append((i, label, tmin, tmax, utype))

    # Also check for transport mode columns (直卡/拖车 in 风筝国际报价)
    transport_cols = []
    for i, v in enumerate(header or []):
        if v is None:
            continue
        s = str(v).strip()
        if s in ("直卡", "拖车", "53'", "26'", "53", "26"):
            transport_cols.append((i, s))

    if not tier_cols and not transport_cols:
        return []

    rules = []
    for r_idx in range(header_idx + 1, len(rows)):
        row = rows[r_idx]
        if not row:
            continue
        dest_val = row[dest_col] if dest_col < len(row) else None
        if not dest_val:
            continue
        dest_str = str(dest_val).strip()
        if not dest_str or dest_str.lower() in ("none", "total", "合计"):
            continue
        dests = expand_destinations(dest_str)
        if not dests:
            continue

        for dest in dests:
            # Standard tier columns
            for col_idx, tier_label, tier_min, tier_max, unit_type in tier_cols:
                if col_idx >= len(row):
                    continue
                base_price, unit_price, ut_override, notes = parse_price_cell(row[col_idx])
                if base_price is None and unit_price is None:
                    if notes:
                        # Store note-only rules for reference
                        rules.append({
                            "destination_warehouse": dest,
                            "pricing_type": "matrix",
                            "transport_mode": None,
                            "pallet_tier_min": tier_min,
                            "pallet_tier_max": tier_max,
                            "unit_type": unit_type,
                            "base_price": 0,
                            "unit_price": 0,
                            "notes": f"[{tier_label}] {notes}",
                            "source_sheet": sheet_name,
                        })
                    continue

                final_ut = ut_override or unit_type
                rules.append({
                    "destination_warehouse": dest,
                    "pricing_type": "matrix",
                    "transport_mode": None,
                    "pallet_tier_min": tier_min,
                    "pallet_tier_max": tier_max,
                    "unit_type": final_ut,
                    "base_price": base_price or 0,
                    "unit_price": unit_price or 0,
                    "notes": notes,
                    "source_sheet": sheet_name,
                })

            # Transport mode columns (风筝国际报价 style)
            for col_idx, mode in transport_cols:
                if col_idx >= len(row):
                    continue
                base_price, unit_price, ut_override, notes = parse_price_cell(row[col_idx])
                if base_price is None and unit_price is None:
                    continue
                rules.append({
                    "destination_warehouse": dest,
                    "pricing_type": "matrix",
                    "transport_mode": mode,
                    "pallet_tier_min": None,
                    "pallet_tier_max": None,
                    "unit_type": ut_override or "per_trip",
                    "base_price": base_price or 0,
                    "unit_price": unit_price or 0,
                    "notes": notes,
                    "source_sheet": sheet_name,
                })

    return rules


def parse_jingdong_sheet(ws, sheet_name):
    """Parse Jing Dong address-based pricing (Address, CITY, 53', 26')."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    # Header is row 0: Address, CITY, 53', 26'
    rules = []
    for r_idx in range(1, len(rows)):
        row = rows[r_idx]
        if not row or len(row) < 3:
            continue
        address = _clean(row[0])
        city = _clean(row[1])
        price_53 = row[2] if len(row) > 2 else None
        price_26 = row[3] if len(row) > 3 else None
        if not city:
            continue

        for price, mode in [(price_53, "53尺"), (price_26, "26尺")]:
            if price is None:
                continue
            try:
                p = float(price)
            except (ValueError, TypeError):
                continue
            rules.append({
                "destination_warehouse": str(city).strip()[:50],
                "pricing_type": "zone",
                "transport_mode": mode,
                "pallet_tier_min": None,
                "pallet_tier_max": None,
                "unit_type": "per_trip",
                "base_price": p,
                "unit_price": 0,
                "notes": str(address)[:200] if address else None,
                "source_sheet": sheet_name,
            })
    return rules


def parse_haptrans_sheet(ws, sheet_name):
    """Parse Haptrans service fee schedule."""
    rows = list(ws.iter_rows(values_only=True))
    rules = []
    for r_idx in range(2, len(rows)):
        row = rows[r_idx]
        if not row:
            continue
        item = _clean(row[0])
        price_str = _clean(row[1])
        desc = _clean(row[2]) if len(row) > 2 else None
        if not item or not price_str:
            continue

        # Try to extract a numeric price
        s = str(price_str)
        m = re.search(r'([\d.]+)', s.replace(",", ""))
        if not m:
            continue

        base_price = float(m.group(1))
        # Check if per-unit pricing
        unit_type = "per_trip"
        unit_price = 0
        if "/板" in s or "/P" in s.upper():
            unit_type = "per_pallet"
            unit_price = base_price
            base_price = 0
        elif "/件" in s or "/箱" in s:
            unit_type = "per_kg"  # approximate: per piece
            unit_price = base_price
            base_price = 0

        rules.append({
            "destination_warehouse": "WAREHOUSE",
            "pricing_type": "service_fee",
            "transport_mode": None,
            "pallet_tier_min": None,
            "pallet_tier_max": None,
            "unit_type": unit_type,
            "base_price": base_price,
            "unit_price": unit_price,
            "notes": f"{item}: {s}" + (f" ({desc})" if desc else ""),
            "source_sheet": sheet_name,
        })
    return rules


def parse_jwapwh_sheet(ws, sheet_name):
    """Parse JWAPWH操作费 weight-tier pricing."""
    rows = list(ws.iter_rows(values_only=True))
    rules = []
    current_section = None
    for r_idx in range(0, len(rows)):
        row = rows[r_idx]
        if not row:
            continue
        first = _clean(row[0])
        if first and ("收费项目" in str(first)):
            continue  # header row

        # Detect section headers like "1、代发订单处理费"
        if first and re.match(r'\d+[、.]', str(first)):
            current_section = str(first)
            continue

        # Look for weight range + price
        weight_range = None
        price = None
        for i, cell in enumerate(row):
            if cell is None:
                continue
            s = str(cell).strip()
            if re.match(r'[\d.]+-[\d.]+kg', s, re.IGNORECASE):
                weight_range = s
            elif re.match(r'>?\s*[\d.]+kg', s, re.IGNORECASE):
                weight_range = s
            elif isinstance(cell, (int, float)):
                price = float(cell)

        if weight_range and price is not None:
            # Parse weight range
            m = re.match(r'([\d.]+)-([\d.]+)', weight_range)
            weight_min = float(m.group(1)) if m else None
            weight_max = float(m.group(2)) if m else None

            rules.append({
                "destination_warehouse": "WAREHOUSE",
                "pricing_type": "weight_tier",
                "transport_mode": None,
                "pallet_tier_min": None,
                "pallet_tier_max": None,
                "weight_min": weight_min,
                "weight_max": weight_max,
                "unit_type": "per_kg",
                "base_price": 0,
                "unit_price": price,
                "notes": current_section,
                "source_sheet": sheet_name,
            })
    return rules


def parse_jason_ml_sheet(ws, sheet_name):
    """Parse Jason_(ML) simple per-pallet pricing. Format: YYZ | 25/P
    Data may start in col 0 or col 1 (sometimes col 0 is empty/merged)."""
    rows = list(ws.iter_rows(values_only=True))
    rules = []
    for r_idx in range(0, len(rows)):
        row = rows[r_idx]
        if not row:
            continue
        # Find first non-None cell as dest, next as price
        vals = [(i, _clean(c)) for i, c in enumerate(row) if _clean(c) is not None]
        if len(vals) < 2:
            continue
        dest = vals[0][1]
        price_str = vals[1][1]
        if not dest or not price_str:
            continue
        # Check if dest looks like a warehouse code
        dests = expand_destinations(str(dest))
        if not dests:
            continue
        s = str(price_str)
        # "25/P" or "100/P" pattern
        m = re.search(r'([\d.]+)\s*/\s*[Pp板]', s)
        if m:
            for d in dests:
                rules.append({
                    "destination_warehouse": d,
                    "pricing_type": "matrix",
                    "transport_mode": None,
                    "pallet_tier_min": 1,
                    "pallet_tier_max": 999,
                    "unit_type": "per_pallet",
                    "base_price": 0,
                    "unit_price": float(m.group(1)),
                    "notes": None,
                    "source_sheet": sheet_name,
                })
            continue
        # Pure number (flat price)
        try:
            price = float(s.replace(",", "").replace("$", ""))
            for d in dests:
                rules.append({
                    "destination_warehouse": d,
                    "pricing_type": "matrix",
                    "transport_mode": None,
                    "pallet_tier_min": 1,
                    "pallet_tier_max": 999,
                    "unit_type": "per_pallet",
                    "base_price": 0,
                    "unit_price": price,
                    "notes": None,
                    "source_sheet": sheet_name,
                })
        except ValueError:
            pass
    return rules


def parse_aseasky_sheet(ws, sheet_name):
    """Parse AseaSky flat/container pricing."""
    rows = list(ws.iter_rows(values_only=True))
    rules = []
    for r_idx in range(0, len(rows)):
        row = rows[r_idx]
        if not row:
            continue
        item = _clean(row[0])
        price = row[1] if len(row) > 1 else None
        if not item or not isinstance(price, (int, float)):
            continue
        notes = _clean(row[2]) if len(row) > 2 else None
        rules.append({
            "destination_warehouse": "ALL",
            "pricing_type": "flat",
            "transport_mode": None,
            "pallet_tier_min": None,
            "pallet_tier_max": None,
            "unit_type": "per_trip",
            "base_price": float(price),
            "unit_price": 0,
            "notes": f"{item}" + (f" ({notes})" if notes else ""),
            "source_sheet": sheet_name,
        })
    return rules


def parse_cbws_sheet(ws, sheet_name):
    """Parse CBWS sheet — notes-based pricing with some scattered data."""
    rows = list(ws.iter_rows(values_only=True))
    rules = []
    # CBWS has sparse data, mostly notes. Look for any structured rows.
    for r_idx in range(0, min(30, len(rows))):
        row = rows[r_idx]
        if not row:
            continue
        vals = [c for c in row if c is not None]
        if not vals:
            continue
        # Look for dest + price patterns
        text = " ".join(str(v) for v in vals)
        # "YYZ系列 35/P" pattern
        dests = re.findall(r'(Y[A-Z]{2}\d?)', text.upper())
        m_price = re.search(r'([\d.]+)/[Pp板]', text)
        if dests and m_price:
            for d in set(dests):
                rules.append({
                    "destination_warehouse": d if d[-1].isdigit() else d,
                    "pricing_type": "matrix",
                    "transport_mode": None,
                    "pallet_tier_min": 1,
                    "pallet_tier_max": 999,
                    "unit_type": "per_pallet",
                    "base_price": 0,
                    "unit_price": float(m_price.group(1)),
                    "notes": text[:200],
                    "source_sheet": sheet_name,
                })
    return rules


# ─── Sheet Router ─────────────────────────────────────────────────────────────

# Sheets that have standard matrix format (仓号 header detected)
MATRIX_SHEETS = [
    'PLD  LT', 'LLL新', 'Panex', 'Focus', 'Ulala', 'Eynex', '鹏鹏',
    'QX', '601， 大榕树，龙形', '风筝国际报价', 'Canada express solution',
    'TanZDowbier', '旺市ESCMDaniel6300地毯Andy', 'JW  挖掘机', '乐速',
    'BTC报价', 'Jason华学斌', 'UBI', '通达内部结算',
]

# Partner name → short_code mapping
PARTNER_CODES = {
    'PLD  LT': ('PLD LT', 'PLD'),
    'LLL新': ('LLL新', 'LLL'),
    'Panex': ('Panex', 'PNX'),
    'Focus': ('Focus', 'FCS'),
    'Ulala': ('Ulala', 'ULA'),
    'Eynex': ('Eynex', 'ENX'),
    '鹏鹏': ('鹏鹏', 'PP'),
    'QX': ('QX', 'QX'),
    '601， 大榕树，龙形': ('601大榕树龙形', '601'),
    '风筝国际报价': ('风筝国际', 'FZ'),
    'Canada express solution': ('Canada Express Solution', 'CES'),
    'TanZDowbier': ('TanZDowbier', 'TZD'),
    '旺市ESCMDaniel6300地毯Andy': ('旺市ESC Daniel', 'ESC'),
    'JW  挖掘机': ('JW挖掘机', 'JW'),
    '乐速': ('乐速', 'LS'),
    'BTC报价': ('BTC', 'BTC'),
    'Jason华学斌': ('Jason华学斌', 'JHX'),
    'UBI': ('UBI', 'UBI'),
    '通达内部结算': ('通达内部结算', 'TD'),
    'Jing Dong(京东）': ('京东 Jing Dong', 'JD'),
    'Haptrans': ('Haptrans/小马物流', 'HPT'),
    'JWAPWH操作费': ('JWAPWH', 'JWAP'),
    'Jason_(ML)14363167 Canada Inc': ('Jason ML', 'JML'),
    'AseaSky': ('AseaSky', 'ASK'),
    'CBWS': ('CBWS', 'CBWS'),
}

# Sheets to skip (image-only, already imported, or empty)
SKIP_SHEETS = [
    '司机工资标准',  # already imported as driver_cost_baselines
    '拾谷',          # empty
    'Rachel',        # empty
    '整合报价卡（更新中）',  # aggregate/mixed, needs manual review
    '郎旭 Langxu',  # contract-based per container, not standard pricing
]

# Image-only sheets (Phase 2.2 — OCR needed)
IMAGE_ONLY_SHEETS = [
    '无忧达卡派', 'LZ', 'ONE', 'Michael', 'Bayou', 'ESC-XIY',
    'Skyworth', 'BTL', 'Sunny Charm', 'Label X', 'TEWM',
    'Igloo+DKA', 'One Express', 'CTCGVTTomXLY', '祁杰',
    '代发客户统一报价表', 'Straightship阿东', '（缘海）YHWL',
    'NoahHammert', 'Terry-Ottawa warehouse', 'Jeff  Don',
]


def ensure_partner(conn, name, short_code):
    """Ensure partner exists, return partner_id."""
    with conn.cursor() as cur:
        cur.execute("SELECT id FROM partners WHERE name = %s", (name,))
        row = cur.fetchone()
        if row:
            return row[0]
        cur.execute(
            "INSERT INTO partners (name, short_code, type, status) VALUES (%s, %s, 'carrier', 'ACTIVE') RETURNING id",
            (name, short_code)
        )
        pid = cur.fetchone()[0]
        conn.commit()
        return pid


def insert_rules(conn, partner_id, rules):
    """Insert rules into partner_pricing_rules. Returns count inserted."""
    count = 0
    with conn.cursor() as cur:
        for r in rules:
            try:
                cur.execute("""
                    INSERT INTO partner_pricing_rules
                      (partner_id, pricing_type, origin_warehouse, destination_warehouse, transport_mode,
                       pallet_tier_min, pallet_tier_max, weight_min, weight_max, unit_type,
                       base_price, unit_price, surcharges, fuel_surcharge_rate,
                       source_sheet, source_type, notes)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    partner_id,
                    r.get("pricing_type", "matrix"),
                    r.get("origin_warehouse"),
                    r["destination_warehouse"],
                    r.get("transport_mode"),
                    r.get("pallet_tier_min"),
                    r.get("pallet_tier_max"),
                    r.get("weight_min"),
                    r.get("weight_max"),
                    r.get("unit_type", "per_trip"),
                    r.get("base_price", 0),
                    r.get("unit_price", 0),
                    json.dumps(r.get("surcharges", {})),
                    r.get("fuel_surcharge_rate"),
                    r.get("source_sheet"),
                    "structured",
                    r.get("notes"),
                ))
                count += 1
            except Exception as e:
                print(f"  [WARN] insert error: {e}")
    conn.commit()
    return count


def main():
    parser = argparse.ArgumentParser(description="导入合作单位报价卡结构化数据")
    parser.add_argument("--preview", action="store_true", help="预览模式，不写入数据库")
    parser.add_argument("--commit", action="store_true", help="写入数据库")
    parser.add_argument("--sheet", type=str, help="只处理指定 sheet")
    args = parser.parse_args()

    if not args.preview and not args.commit:
        print("请指定 --preview 或 --commit")
        return 1

    if not os.path.isfile(FILE_QUOTE):
        print(f"文件不存在: {FILE_QUOTE}")
        return 1

    print(f"读取 {FILE_QUOTE} ...")
    wb = openpyxl.load_workbook(FILE_QUOTE, data_only=True)
    all_sheets = wb.sheetnames
    print(f"共 {len(all_sheets)} 个 sheet\n")

    conn = None
    if args.commit:
        url = _load_database_url()
        if not url or "..." in url:
            print("错误: 未找到有效的 DATABASE_URL")
            return 1
        conn = psycopg2.connect(url)

    total_rules = 0
    total_partners = 0
    summary = []

    for sn in all_sheets:
        if args.sheet and sn != args.sheet:
            continue
        if sn in SKIP_SHEETS:
            summary.append(f"  SKIP  {sn} (跳过)")
            continue
        if sn in IMAGE_ONLY_SHEETS:
            summary.append(f"  IMAGE {sn} (需 OCR，Phase 2.2)")
            continue

        ws = wb[sn]
        rules = []

        # Route to appropriate parser
        if sn in MATRIX_SHEETS:
            rules = parse_matrix_sheet(ws, sn)
        elif sn == 'Jing Dong(京东）':
            rules = parse_jingdong_sheet(ws, sn)
        elif sn == 'Haptrans':
            rules = parse_haptrans_sheet(ws, sn)
        elif sn == 'JWAPWH操作费':
            rules = parse_jwapwh_sheet(ws, sn)
        elif sn.startswith('Jason_(ML)'):
            rules = parse_jason_ml_sheet(ws, sn)
        elif sn == 'AseaSky':
            rules = parse_aseasky_sheet(ws, sn)
        elif sn == 'CBWS':
            rules = parse_cbws_sheet(ws, sn)
        else:
            # Try matrix parser as fallback
            rules = parse_matrix_sheet(ws, sn)
            if not rules:
                summary.append(f"  ???   {sn} (未识别格式，0 rules)")
                continue

        if not rules:
            summary.append(f"  EMPTY {sn} (解析到 0 条规则)")
            continue

        # Get partner name/code
        partner_name, partner_code = PARTNER_CODES.get(sn, (sn, sn[:5].upper()))

        # Filter out note-only rules (base_price=0 and unit_price=0 and has notes)
        price_rules = [r for r in rules if r.get("base_price", 0) > 0 or r.get("unit_price", 0) > 0]
        note_rules = [r for r in rules if r.get("base_price", 0) == 0 and r.get("unit_price", 0) == 0]

        status = f"  OK    {sn} → {partner_name} [{partner_code}]: {len(price_rules)} price rules"
        if note_rules:
            status += f" + {len(note_rules)} notes"

        # Show sample rules in preview
        if args.preview and price_rules:
            status += "\n"
            for r in price_rules[:5]:
                dest = r['destination_warehouse']
                tier = f"{r.get('pallet_tier_min','')}-{r.get('pallet_tier_max','')}" if r.get('pallet_tier_min') else "all"
                mode = r.get('transport_mode') or ''
                bp = r.get('base_price', 0)
                up = r.get('unit_price', 0)
                ut = r.get('unit_type', '')
                line = f"          {dest:8s} tier={tier:7s} {mode:6s} base=${bp:.0f} unit=${up:.0f} ({ut})"
                if r.get('notes'):
                    line += f" [{r['notes'][:40]}]"
                status += line + "\n"
            if len(price_rules) > 5:
                status += f"          ... and {len(price_rules) - 5} more\n"

        summary.append(status)
        total_rules += len(price_rules)
        total_partners += 1

        # Commit to DB
        if args.commit and conn and price_rules:
            partner_id = ensure_partner(conn, partner_name, partner_code)
            inserted = insert_rules(conn, partner_id, price_rules)
            summary[-1] += f" → DB: {inserted} inserted"

    wb.close()

    print("=" * 70)
    print("合作单位报价卡 导入摘要")
    print("=" * 70)
    for line in summary:
        print(line)
    print("-" * 70)
    print(f"合计: {total_partners} 个 partner, {total_rules} 条定价规则")
    if args.preview:
        print("\n这是预览模式。确认无误后请执行: python scripts/import-partner-pricing.py --commit")

    if conn:
        conn.close()
    return 0


if __name__ == "__main__":
    exit(main())
