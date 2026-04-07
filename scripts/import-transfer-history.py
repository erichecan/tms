#!/usr/bin/env python3
"""
Phase 4: 转运汇总表历史数据导入
用法:
  python3 scripts/import-transfer-history.py --preview
  python3 scripts/import-transfer-history.py --commit
  python3 scripts/import-transfer-history.py --commit --sheet "BTL"
"""
import sys, os, re, argparse, datetime
from pathlib import Path

# Load .env
env_path = Path(__file__).parent.parent / 'apps' / 'backend' / '.env'
if env_path.exists():
    for line in env_path.read_text().splitlines():
        line = line.strip()
        if line and not line.startswith('#') and '=' in line:
            k, v = line.split('=', 1)
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))

import openpyxl
import psycopg2
from psycopg2.extras import execute_values

EXCEL_FILE = '2026年转运汇总表.xlsx'
DATABASE_URL = os.environ.get('DATABASE_URL', '')

# ── Sheet → partner name 映射 ─────────────────────────────────────────────────
# 值为 DB 中 partners.name，None 表示需从行内数据动态解析
SHEET_PARTNER_MAP = {
    'EynexCAME':              'Eynex',
    'Straightship Toronto':   'Straightship阿东',
    'Straightship Calgary':   'Straightship阿东',
    'Straightship Vancouver': 'Straightship阿东',
    'XLY（徐）':              'CTC/GVT/Tom/XLY',
    'BTL':                    'BTL',
    'ASS柜子':                'ASS柜子',
    '其他转运':               None,        # 行内 col0 有 partner
    '第三方运输':             None,
    'Bisonex-祁杰':           '祁杰',
    '鹏鹏':                   '鹏鹏',
    'GVT柜子':                'CTC/GVT/Tom/XLY',
    'Skyworth':               'Skyworth',
    'ESC-XIY':                'ESC-XIY',
    'Focus':                  'Focus',
    'LabelX':                 'Label X',
    'Hammert':                'NoahHammert',
    '2025转运已完成':         None,        # col0 动态
    '2025第三方-已完成':      None,        # 结构不规则，跳过
    '阿东（3600A）':          'Straightship阿东',
    '2024已完成':             None,        # col1 动态 (col0=序号)
    'YH已完成2025':           None,        # col0 动态
}

# ── 特殊 sheet 结构覆盖 ───────────────────────────────────────────────────────
# skip_header_rows: 跳过前N行作为header
# col_map: 显式列映射 {field: col_index}
# partner_col: partner名所在列index（动态partner时使用）
SHEET_OVERRIDES = {
    '2024已完成': {
        'skip_header_rows': 2,    # 行1+行2是双行表头，数据从行3开始
        'partner_col': 1,         # col0=序号, col1=客户名
    },
    '2025转运已完成': {
        'skip_header_rows': 0,    # 无表头行，直接是数据
        'col_map': {
            'container_no': 1,
            'entry_method': 3,
            'arrival_date': 4,
            'fba_shipment_id': 7,
            'sku': 8,
            'piece_count': 9,
            'cbm': 10,
            'dest_warehouse': 11,
            'delivery_type': 12,
            'pallet_count': 13,
            'appt1': 14,
            'appt2': 15,
            'notes': 17,
        },
        'partner_col': 0,
    },
    'YH已完成2025': {
        'skip_header_rows': 0,
        'col_map': {
            'container_no': 1,
            'entry_method': 3,
            'arrival_date': 4,
            'fba_shipment_id': 7,
            'sku': 8,
            'piece_count': 9,
            'cbm': 10,
            'dest_warehouse': 11,
            'delivery_type': 12,
            'pallet_count': 13,
            'appt1': 14,
            'appt2': 15,
        },
        'partner_col': 0,
    },
    '2025第三方-已完成': {
        'skip': True,             # 结构过于不规则，跳过
    },
}

# ── 列名别名归一 ──────────────────────────────────────────────────────────────
COL_ALIASES = {
    '柜号': 'container_no',
    '所在仓库': 'warehouse',
    '进仓方式': 'entry_method',
    '到仓日': 'arrival_date',
    '卸柜情况': 'unload_note',
    '分货SKU': 'sku',
    'FBA': 'fba_shipment_id',
    'POLIS': 'po_list', 'POLIST': 'po_list', 'POLIST ': 'po_list',
    'PoList': 'po_list', 'PO LIST': 'po_list', 'POLIST': 'po_list',
    '件/箱数': 'piece_count',
    '件数': 'piece_count',
    '方数': 'cbm',
    '目的仓': 'dest_warehouse',
    '派送指令': 'delivery_type',
    '托数': 'pallet_count',
    '预约时间1': 'appt1',
    '预约时间2': 'appt2',
    '预约时间3': 'appt3',
    '备注': 'notes',
}

def normalize_col(name):
    if name is None:
        return None
    s = str(name).strip()
    return COL_ALIASES.get(s, s.lower().replace(' ', '_'))

def parse_header(ws, header_row_num=1):
    """返回 {field: col_index} 及 partner_col_name (首列若含合作单位名则记录)"""
    header_row = list(ws.iter_rows(min_row=header_row_num, max_row=header_row_num, values_only=True))[0]
    col_map = {}
    first_col_val = str(header_row[0]).strip() if header_row[0] else ''
    # 首列是否是 partner 名称列（非字段名）
    is_partner_col = first_col_val not in ('柜号', '客户', '', 'None') and \
                     first_col_val not in COL_ALIASES

    for i, val in enumerate(header_row):
        norm = normalize_col(val)
        if norm and norm != 'none':
            col_map[norm] = i

    return col_map, (first_col_val if is_partner_col else None)

def safe_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s and s.lower() != 'none' else None

def safe_date(v):
    if isinstance(v, (datetime.date, datetime.datetime)):
        return v.date() if isinstance(v, datetime.datetime) else v
    if v is None:
        return None
    s = str(v).strip()
    if not s or s.lower() == 'none':
        return None
    for fmt in ('%Y/%m/%d', '%Y-%m-%d', '%d/%m/%Y'):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None

def safe_num(v):
    if v is None:
        return None
    try:
        f = float(str(v).strip())
        return f if f else None
    except (ValueError, TypeError):
        return None

def parse_pallet(v):
    """托数字段可能是 '5P', '3P+2LP', '1P', int, float"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(v) if v else None
    s = str(v).strip()
    # Extract leading integer
    m = re.match(r'^(\d+)', s)
    if m:
        return int(m.group(1))
    return None

def generate_order_no(sheet_short, arrival_date, seq):
    d = arrival_date.strftime('%Y%m%d') if arrival_date else '00000000'
    return f"TO-{sheet_short[:3].upper()}-{d}-{seq:04d}"

# ── 解析一个 sheet 的所有记录 ──────────────────────────────────────────────────
def parse_sheet(ws, sheet_name, fixed_partner_name):
    """
    返回 list of {order_fields, lines:[line_fields]}
    """
    override = SHEET_OVERRIDES.get(sheet_name, {})
    skip_rows = override.get('skip_header_rows', None)
    explicit_col_map = override.get('col_map', None)
    explicit_partner_col = override.get('partner_col', None)

    if explicit_col_map is not None:
        # 无表头 sheet：直接用显式列映射
        col_map = explicit_col_map
        dyn_partner_col_header = '__explicit__' if explicit_partner_col is not None else None
        data_start_row = 2  # default
    elif skip_rows is not None and skip_rows > 1:
        # 多行表头：以最后一个表头行解析
        col_map, dyn_partner_col_header = parse_header(ws, header_row_num=skip_rows)
        data_start_row = skip_rows + 1
    else:
        col_map, dyn_partner_col_header = parse_header(ws)
        data_start_row = 2

    # 确定关键列 index
    def ci(field):
        return col_map.get(field)

    # 迭代数据行（跳过 header 第1行）
    orders = []
    current_order = None
    current_lines = []

    def flush():
        nonlocal current_order, current_lines
        if current_order is not None:
            orders.append({'order': current_order, 'lines': current_lines})
        current_order = None
        current_lines = []

    all_rows = list(ws.iter_rows(min_row=data_start_row, values_only=True))

    for row in all_rows:
        # 跳过全空行
        if all(v is None for v in row):
            continue

        # 检测是否是新柜号行（container_no 列非空）
        container_val = None
        if ci('container_no') is not None:
            container_val = safe_str(row[ci('container_no')])
        # 有些 sheet 柜号在 col 0 或 col 1
        if container_val is None and len(row) > 1:
            for cname in ('container_no',):
                idx = ci(cname)
                if idx is not None:
                    container_val = safe_str(row[idx])
                    break

        is_new_container = container_val and container_val.lower() != 'none'

        if is_new_container:
            flush()
            # 动态 partner
            partner_name = fixed_partner_name
            if partner_name is None:
                if explicit_partner_col is not None and len(row) > explicit_partner_col:
                    partner_name = safe_str(row[explicit_partner_col])
                elif dyn_partner_col_header is None:
                    # Try col 0
                    partner_name = safe_str(row[0]) if len(row) > 0 else None

            arrival = safe_date(row[ci('arrival_date')]) if ci('arrival_date') is not None else None
            warehouse = safe_str(row[ci('warehouse')]) if ci('warehouse') is not None else None
            entry_method = safe_str(row[ci('entry_method')]) if ci('entry_method') is not None else None
            notes_val = safe_str(row[ci('notes')]) if ci('notes') is not None else None

            current_order = {
                'container_no': container_val,
                'partner_name': partner_name,
                'warehouse': warehouse,
                'entry_method': entry_method,
                'arrival_date': arrival,
                'source_sheet': sheet_name,
                'notes': notes_val,
            }
            current_lines = []

        # Append line regardless (new or continuation)
        if current_order is not None:
            fba = safe_str(row[ci('fba_shipment_id')]) if ci('fba_shipment_id') is not None else None
            sku = safe_str(row[ci('sku')]) if ci('sku') is not None else None
            po  = safe_str(row[ci('po_list')]) if ci('po_list') is not None else None
            dest = safe_str(row[ci('dest_warehouse')]) if ci('dest_warehouse') is not None else None
            if dest and len(dest) > 20:
                dest = dest[:20]
            pallets = parse_pallet(row[ci('pallet_count')]) if ci('pallet_count') is not None else None
            pieces  = safe_num(row[ci('piece_count')]) if ci('piece_count') is not None else None
            cbm     = safe_num(row[ci('cbm')]) if ci('cbm') is not None else None
            delivery = safe_str(row[ci('delivery_type')]) if ci('delivery_type') is not None else None
            appt1   = safe_str(row[ci('appt1')]) if ci('appt1') is not None else None
            appt2   = safe_str(row[ci('appt2')]) if ci('appt2') is not None else None
            appt3   = safe_str(row[ci('appt3')]) if ci('appt3') is not None else None
            line_notes = safe_str(row[ci('notes')]) if ci('notes') is not None else None

            # Only add line if has meaningful data
            if any([fba, sku, po, dest, pallets, pieces]):
                current_lines.append({
                    'fba_shipment_id': fba,
                    'sku': sku,
                    'po_list': po,
                    'dest_warehouse': dest,
                    'pallet_count': pallets,
                    'piece_count': int(pieces) if pieces else None,
                    'cbm': cbm,
                    'delivery_type': delivery,
                    'appt1': appt1,
                    'appt2': appt2,
                    'appt3': appt3,
                    'notes': line_notes,
                })

    flush()
    return orders

# ── 主逻辑 ────────────────────────────────────────────────────────────────────
def get_partner_id_map(conn):
    with conn.cursor() as cur:
        cur.execute("SELECT id, name, short_code FROM partners")
        rows = cur.fetchall()
    mapping = {}
    for pid, name, code in rows:
        mapping[name] = pid
        if code:
            mapping[code] = pid
    return mapping

def match_partner(name, pid_map):
    if not name:
        return None
    if name in pid_map:
        return pid_map[name]
    # Fuzzy: check if name is substring of any key
    name_lower = name.lower()
    for k, v in pid_map.items():
        if name_lower in k.lower() or k.lower() in name_lower:
            return v
    return None

def run(args):
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    sheets_to_process = [args.sheet] if args.sheet else list(SHEET_PARTNER_MAP.keys())

    conn = None
    pid_map = {}
    if args.commit or args.preview:
        conn = psycopg2.connect(DATABASE_URL)
        pid_map = get_partner_id_map(conn)

    grand_orders = 0
    grand_lines = 0
    unmatched_partners = set()

    for sheet_name in sheets_to_process:
        if sheet_name not in wb.sheetnames:
            print(f"[SKIP] Sheet not found: {sheet_name}")
            continue

        override = SHEET_OVERRIDES.get(sheet_name, {})
        if override.get('skip'):
            print(f"[{sheet_name}] 跳过（结构不规则）")
            continue

        ws = wb[sheet_name]
        fixed_partner = SHEET_PARTNER_MAP.get(sheet_name)  # May be None
        records = parse_sheet(ws, sheet_name, fixed_partner)

        # Stats
        partner_hit = 0
        partner_miss = 0
        unmatched_in_sheet = set()
        for rec in records:
            pname = rec['order'].get('partner_name')
            pid = match_partner(pname, pid_map)
            if pid:
                partner_hit += 1
            else:
                partner_miss += 1
                if pname:
                    unmatched_in_sheet.add(pname)
                    unmatched_partners.add(pname)

        total_lines = sum(len(r['lines']) for r in records)
        match_pct = f"{100*partner_hit//len(records)}%" if records else "N/A"
        print(f"[{sheet_name}] {len(records)} 柜 / {total_lines} 行  partner匹配率:{match_pct}", end='')
        if unmatched_in_sheet:
            print(f"  未匹配:{sorted(unmatched_in_sheet)}", end='')
        print()

        grand_orders += len(records)
        grand_lines += total_lines

        if args.commit:
            inserted_orders = 0
            inserted_lines = 0
            seq = 1
            with conn.cursor() as cur:
                for rec in records:
                    o = rec['order']
                    pname = o.get('partner_name')
                    pid = match_partner(pname, pid_map)
                    order_no = generate_order_no(sheet_name, o['arrival_date'], seq)
                    seq += 1

                    def trunc(v, n):
                        return v[:n] if v and len(v) > n else v

                    # Upsert transfer_order
                    cur.execute("""
                        INSERT INTO transfer_orders
                          (order_no, partner_id, container_no, warehouse, entry_method,
                           arrival_date, notes, status)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,'COMPLETED')
                        ON CONFLICT (order_no) DO NOTHING
                        RETURNING id
                    """, (order_no, pid, trunc(o['container_no'], 30), trunc(o['warehouse'], 50),
                          trunc(o['entry_method'], 30), o['arrival_date'], o['notes']))
                    row = cur.fetchone()
                    if not row:
                        continue  # duplicate
                    order_id = row[0]
                    inserted_orders += 1

                    # Insert lines
                    for ln_no, ln in enumerate(rec['lines'], start=1):
                        cur.execute("""
                            INSERT INTO transfer_order_lines
                              (transfer_order_id, line_no, fba_shipment_id, sku, po_list,
                               dest_warehouse, pallet_count, piece_count, cbm,
                               delivery_type, notes)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        """, (order_id, ln_no, trunc(ln['fba_shipment_id'], 50), trunc(ln['sku'], 100), ln['po_list'],
                              trunc(ln['dest_warehouse'], 20), ln['pallet_count'], ln['piece_count'], ln['cbm'],
                              trunc(ln['delivery_type'], 30), ln['notes']))
                        inserted_lines += 1

            conn.commit()
            print(f"  → 已写入 {inserted_orders} 转运单 / {inserted_lines} 明细行")

    print(f"\n汇总: {grand_orders} 转运单 / {grand_lines} 明细行")
    if unmatched_partners:
        print(f"未匹配合作单位: {sorted(unmatched_partners)}")
        print("  提示: 可在 partners 表手动新增上述合作单位后重新运行")

    if conn:
        conn.close()

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--preview', action='store_true', help='预览（不写库）')
    parser.add_argument('--commit', action='store_true', help='写入数据库')
    parser.add_argument('--sheet', type=str, default=None, help='只处理指定 sheet')
    args = parser.parse_args()

    if not args.preview and not args.commit:
        parser.print_help()
        sys.exit(1)

    run(args)
