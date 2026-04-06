#!/usr/bin/env python3
# 2026-03-13: 将根目录「2026年转运汇总表.xlsx」「合作单位报价卡.xlsx」导入生产库
# 用法一: DATABASE_URL='postgresql://user:pass@host/db?sslmode=require' python scripts/import_excel_to_prod.py
# 用法二: 在 apps/backend/.env 中配置 DATABASE_URL 后直接运行 python scripts/import_excel_to_prod.py
# 依赖: pip install openpyxl psycopg2-binary
# OCR 依赖（可选，用于识别图片报价）: pip install pytesseract Pillow
#   并安装系统 Tesseract: brew install tesseract tesseract-lang  (macOS)
#                          apt-get install tesseract-ocr tesseract-ocr-chi-sim  (Ubuntu)

import os
import re
import uuid
import json
import base64
from datetime import datetime
from decimal import Decimal
from io import BytesIO

import openpyxl
import psycopg2
from psycopg2.extras import execute_values

# OCR 可选依赖
try:
    import pytesseract
    from PIL import Image as PILImage
    _OCR_AVAILABLE = True
except ImportError:
    _OCR_AVAILABLE = False

# 环境变量可覆盖 Tesseract 语言包配置，默认简体中文+英文
_OCR_LANG = os.environ.get("OCR_LANG", "chi_sim+eng")

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ENV_PATH = os.path.join(ROOT, "apps", "backend", ".env")


def _load_database_url():
    """从环境变量或 apps/backend/.env 读取 DATABASE_URL；若为占位符则尝试 .env。"""
    url = os.environ.get("DATABASE_URL", "").strip()
    # 占位符（如文档里的 ...）或未设置时，从 .env 读取
    if not url or "..." in url or url == "postgresql://":
        if os.path.isfile(ENV_PATH):
            with open(ENV_PATH, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if line.startswith("DATABASE_URL="):
                        # 支持 DATABASE_URL="..." 或 DATABASE_URL=...
                        val = line.split("=", 1)[1].strip()
                        if val.startswith('"') and val.endswith('"'):
                            val = val[1:-1]
                        elif val.startswith("'") and val.endswith("'"):
                            val = val[1:-1]
                        url = val.strip()
                        break
    return url.strip() if url else ""


FILE_TRANSFER = os.path.join(ROOT, "2026年转运汇总表.xlsx")
FILE_QUOTE = os.path.join(ROOT, "合作单位报价卡.xlsx")

# 中文表头 -> 英文 key（转运表）
TRANSFER_HEADERS = {
    "柜号": "container_no",
    "所在仓库": "warehouse_id",
    "进仓方式": "entry_method",
    "到仓日": "arrival_date",
    "卸柜情况": "unload_status",
    "分货SKU": "sku",
    "FBA": "fba_shipment_id",
    "PoList": "po_list",
    "件/箱数": "piece_count",
    "方数": "cbm",
    "目的仓": "dest_warehouse",
    "派送指令": "delivery_instruction",
    "托数": "pallet_count",
    "预约时间1": "appt1",
    "预约时间2": "appt2",
    "预约时间3": "appt3",
    "备注": "notes",
    "客户": "customer",
}


def _clean(v):
    if v is None:
        return None
    if isinstance(v, (datetime,)):
        return v.date() if hasattr(v, "date") else v
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        if v == int(v):
            return int(v)
        return float(v)
    s = str(v).strip()
    return s if s else None


def _col_index(row, name):
    for i, cell in enumerate(row):
        if cell is None:
            continue
        s = str(cell).strip()
        if s == name or name in s:
            return i
    return -1


def find_header_row(rows):
    for r_idx, row in enumerate(rows):
        if _col_index(row, "柜号") >= 0:
            return r_idx, row
        if _col_index(row, "客户") >= 0 and _col_index(row, "柜号") >= 0:
            return r_idx, row
    return -1, None


def parse_transfer_sheet(ws):
    """解析转运表的一个 sheet，返回 (containers_list, container_items_list)。"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    hr, header_row = find_header_row(rows)
    if hr < 0:
        return [], []
    # 建立列名 -> 列下标
    col_map = {}
    for i, cell in enumerate(header_row):
        if cell is None:
            continue
        s = str(cell).strip()
        for cn, en in TRANSFER_HEADERS.items():
            if cn in s or s == cn:
                col_map[en] = i
                break
    if "container_no" not in col_map and "customer" in col_map:
        # 有的 sheet 柜号在第二列
        for i, cell in enumerate(header_row):
            if cell is None:
                continue
            if str(cell).strip() == "柜号":
                col_map["container_no"] = i
                break
    containers = []
    container_items = []
    current_container = None
    sheet_name = ws.title
    for r_idx in range(hr + 1, len(rows)):
        row = rows[r_idx]
        if not row:
            continue
        try:
            row = tuple(row) if not isinstance(row, (tuple, list)) else row
        except Exception:
            continue

        def get(k):
            idx = col_map.get(k, -1)
            if idx < 0 or idx >= len(row):
                return None
            v = row[idx]
            return _clean(v)
        container_no = get("container_no")
        customer = get("customer")
        # 新柜：有柜号或首列有客户名（部分 sheet 客户+柜号在同一行）
        if container_no or (customer and not current_container):
            if container_no:
                cid = "CTR-" + str(uuid.uuid4())[:8].upper()
                arrival = get("arrival_date")
                if isinstance(arrival, datetime):
                    arrival = arrival.date()
                elif arrival is not None and not hasattr(arrival, "year"):
                    arrival = None
                containers.append({
                    "id": cid,
                    "container_no": str(container_no or ("UNKNOWN-" + cid))[:50],
                    "warehouse_id": str(get("warehouse_id") or "")[:20] if get("warehouse_id") is not None else None,
                    "entry_method": str(get("entry_method") or "")[:20],
                    "arrival_date": arrival,
                    "unload_status": str(get("unload_status") or "")[:200] if get("unload_status") is not None else None,
                    "customer_id": None,  # 后面统一用 sheet 对应客户
                    "notes": str(get("notes") or "")[:500] if get("notes") is not None else None,
                    "sheet": sheet_name,
                    "customer_raw": customer,
                })
                current_container = cid
        # 当前行可能是 item（有 FBA 或 SKU 或 件数）
        if current_container:
            fba = get("fba_shipment_id")
            sku = get("sku")
            pieces = get("piece_count")
            if fba or sku or (pieces is not None and str(pieces).strip()):
                item_id = "CI-" + str(uuid.uuid4())[:8].upper()
                cbm = get("cbm")
                if cbm is not None and isinstance(cbm, str) and "/" in cbm:
                    cbm = None
                try:
                    pc_int = int(float(pieces)) if pieces is not None else None
                except (TypeError, ValueError):
                    pc_int = None
                pallet = get("pallet_count")
                if pallet is not None:
                    pallet = str(pallet).strip()[:10]
                container_items.append({
                    "id": item_id,
                    "container_id": current_container,
                    "sku": (str(sku)[:100] if sku is not None else None),
                    "fba_shipment_id": (str(fba)[:50] if fba is not None else None),
                    "po_list": (str(get("po_list") or "")[:50] if get("po_list") is not None else None),
                    "piece_count": pc_int,
                    "cbm": float(cbm) if cbm is not None and isinstance(cbm, (int, float)) else None,
                    "dest_warehouse": (str(get("dest_warehouse") or "")[:20] if get("dest_warehouse") is not None else None),
                    "pallet_count": pallet,
                    "notes": str(get("notes") or "")[:500] if get("notes") is not None else None,
                    "appt1": str(get("appt1") or "")[:500] if get("appt1") is not None else None,
                })
    return containers, container_items


def ensure_customer(conn, customer_id, name, details=None):
    try:
        with conn.cursor() as cur:
            cur.execute(
                """INSERT INTO customers (id, name, email, phone, businessType, status, details) 
                   VALUES (%s, %s, %s, %s, %s, %s, %s) 
                   ON CONFLICT (id) DO UPDATE SET 
                     name = EXCLUDED.name,
                     details = COALESCE(customers.details, '{}'::jsonb) || EXCLUDED.details
                """,
                (customer_id, name or customer_id, "", "", "STANDARD", "ACTIVE", json.dumps(details) if details else '{}')
            )
        conn.commit()
    except Exception as e:
        conn.rollback()
        print("  [WARN] customer", customer_id, e)


def slug(s):
    s = re.sub(r"[^\w\u4e00-\u9fff]", "", str(s))
    return (s[:15] or "C") + "-" + str(uuid.uuid4())[:4].upper()


def import_transfer(conn, filepath):
    if not os.path.isfile(filepath):
        print("文件不存在:", filepath)
        return 0, 0
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    all_containers = []
    all_items = []
    sheet_to_customer = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        if not sn or sn.startswith("2024") or sn.startswith("2025第三方") or "已完成" in sn:
            continue
        try:
            c_list, i_list = parse_transfer_sheet(ws)
        except Exception as e:
            print("  [WARN] sheet", sn, "parse error:", e)
            continue
        if not c_list and not i_list:
            continue
        # 用 sheet 名作为客户 id 的由来
        cust_id = "C-" + re.sub(r"[^\w\u4e00-\u9fff]", "", sn)[:12] or "SHEET"
        if cust_id not in sheet_to_customer:
            sheet_to_customer[cust_id] = sn
        for c in c_list:
            c["customer_id"] = cust_id
            all_containers.append(c)
        all_items.extend(i_list)
    wb.close()

    # 确保客户存在
    for cid, cname in sheet_to_customer.items():
        ensure_customer(conn, cid, cname)

    inserted_c = 0
    inserted_i = 0
    with conn.cursor() as cur:
        for c in all_containers:
            try:
                cur.execute("""
                    INSERT INTO containers (id, container_no, warehouse_id, entry_method, arrival_date, customer_id, notes, status)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, 'NEW')
                    ON CONFLICT (id) DO NOTHING
                """, (
                    c["id"], c["container_no"], c.get("warehouse_id"), c.get("entry_method"),
                    c.get("arrival_date"), c["customer_id"], c.get("notes")
                ))
                if cur.rowcount:
                    inserted_c += 1
            except Exception as e:
                print("  [WARN] container insert", c.get("container_no"), e)
        for i in all_items:
            try:
                cur.execute("""
                    INSERT INTO container_items (id, container_id, sku, fba_shipment_id, po_list, piece_count, cbm, dest_warehouse, pallet_count, notes, status)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'PENDING')
                    ON CONFLICT (id) DO NOTHING
                """, (
                    i["id"], i["container_id"], i.get("sku"), i.get("fba_shipment_id"), i.get("po_list"),
                    i.get("piece_count"), i.get("cbm"), i.get("dest_warehouse"), i.get("pallet_count"), i.get("notes")
                ))
                if cur.rowcount:
                    inserted_i += 1
            except Exception as e:
                print("  [WARN] item insert", i.get("id"), e)
    conn.commit()
    return inserted_c, inserted_i


def parse_quote_sheet(ws):
    """解析报价卡 sheet，返回 list of dict: customer_id, destination_code, vehicle_type, pallet_tier, base_price, per_pallet_price。"""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 4:
        return []
    # 找「仓号」行
    header_row_idx = -1
    for r in range(min(5, len(rows))):
        row = rows[r]
        for c, v in enumerate(row or []):
            if v and "仓号" in str(v):
                header_row_idx = r
                break
        if header_row_idx >= 0:
            break
    if header_row_idx < 0:
        return []
    header = rows[header_row_idx]
    # 仓号列
    dest_col = -1
    for i, v in enumerate(header or []):
        if v and "仓号" in str(v):
            dest_col = i
            break
    if dest_col < 0:
        return []
    # 价格列：1-4板 等
    tier_cols = []
    for i, v in enumerate(header or []):
        if v is None:
            continue
        s = str(v).strip()
        if "1-4" in s or "1-3" in s:
            tier_cols.append((i, "1-4板"))
        elif "5-8" in s or "5-14" in s:
            tier_cols.append((i, "5-14板"))
        elif "9-13" in s:
            tier_cols.append((i, "9-13板"))
        elif "15-28" in s or "14-28" in s:
            tier_cols.append((i, "15-28板"))
        elif "散" in s or "拼" in s:
            tier_cols.append((i, "散板"))
    if not tier_cols:
        return []
    results = []
    for r in range(header_row_idx + 1, min(header_row_idx + 10, len(rows))):
        row = rows[r]
        if not row:
            continue
        dest = _clean(row[dest_col] if dest_col < len(row) else None)
        if not dest or len(str(dest)) < 2:
            continue
        # YYZ3/YYZ4 -> 拆成多个 destination
        dest_codes = re.findall(r"Y[A-Z]{2}\d+|YOO\d|YHM\d|YYC\d|YEG\d|YVR\d|YYZ\d", str(dest).upper())
        if not dest_codes:
            dest_codes = [str(dest).strip()[:10]]
        for dc in dest_codes[:3]:
            for col_idx, tier in tier_cols:
                if col_idx >= len(row):
                    continue
                v = row[col_idx]
                if v is None:
                    continue
                if isinstance(v, (int, float)):
                    results.append({
                        "destination_code": dc[:10],
                        "pallet_tier": tier,
                        "base_price": float(v),
                        "per_pallet_price": None,
                    })
                else:
                    s = str(v).strip()
                    if not s:
                        continue
                    # "10/板" -> per_pallet=10
                    m = re.search(r"([\d.]+)\s*/\s*板", s)
                    if m:
                        results.append({
                            "destination_code": dc[:10],
                            "pallet_tier": tier,
                            "base_price": None,
                            "per_pallet_price": float(m.group(1)),
                        })
                    else:
                        try:
                            results.append({
                                "destination_code": dc[:10],
                                "pallet_tier": tier,
                                "base_price": float(s.replace(",", "").replace(" ", "")),
                                "per_pallet_price": None,
                            })
                        except ValueError:
                            pass
    return results


def _check_ocr_available():
    """检查 OCR 依赖是否可用，返回 (ok: bool, error_msg: str)。"""
    if not _OCR_AVAILABLE:
        return False, (
            "缺少 OCR 依赖，无法识别图片报价。\n"
            "  请安装: pip install pytesseract Pillow\n"
            "  并安装 Tesseract: brew install tesseract tesseract-lang  (macOS)\n"
            "                     apt-get install tesseract-ocr tesseract-ocr-chi-sim  (Ubuntu)"
        )
    try:
        version = pytesseract.get_tesseract_version()
        return True, f"Tesseract {version}"
    except pytesseract.TesseractNotFoundError:
        return False, (
            "未找到 Tesseract 可执行文件。\n"
            "  macOS: brew install tesseract tesseract-lang\n"
            "  Ubuntu: apt-get install tesseract-ocr tesseract-ocr-chi-sim\n"
            "  或设置环境变量 TESSDATA_PREFIX 指向语言包目录。"
        )
    except Exception as e:
        return False, f"Tesseract 初始化失败: {e}"


def _ocr_image_bytes(img_bytes):
    """对图片二进制数据执行 OCR，返回识别文本；失败返回 None。"""
    try:
        img = PILImage.open(BytesIO(img_bytes))
        # 转 RGB 确保兼容性
        if img.mode not in ("RGB", "L"):
            img = img.convert("RGB")
        text = pytesseract.image_to_string(img, lang=_OCR_LANG)
        return text
    except Exception as e:
        print(f"  [WARN] OCR 识别失败: {e}")
        return None


# 板数档的识别规则：正则模式 -> 标准化名称（与 parse_quote_sheet 保持一致）
_TIER_PATTERNS = [
    (re.compile(r"1[-–~]4"), "1-4板"),
    (re.compile(r"1[-–~]3"), "1-4板"),
    (re.compile(r"5[-–~]14"), "5-14板"),
    (re.compile(r"5[-–~]8"), "5-14板"),
    (re.compile(r"9[-–~]13"), "9-13板"),
    (re.compile(r"15[-–~]28"), "15-28板"),
    (re.compile(r"14[-–~]28"), "15-28板"),
    (re.compile(r"散[板货]?"), "散板"),
    (re.compile(r"拼[板货]?"), "散板"),
]

# 仓号正则：匹配 YYZ3, YHM1, YOO1, YVR1, YYC1, YEG1 等
_FC_CODE_RE = re.compile(r"\b(Y[A-Z]{2}\d+|YOO\d|YHM\d)\b", re.IGNORECASE)

# 金额正则：匹配如 1200、1,200、1200.00 等，不匹配日期或代码里的数字
_PRICE_RE = re.compile(r"(?<!\w)(\d{3,6}(?:[,，]\d{3})*(?:\.\d{1,2})?|\d{2,6}\.\d{2})(?!\w)")


def _parse_tier_from_text(text):
    """从文本中识别板数档名称，返回名称或 None。"""
    for pat, name in _TIER_PATTERNS:
        if pat.search(text):
            return name
    return None


def _extract_prices(text):
    """从文本片段中提取所有金额数值，返回 float 列表。"""
    results = []
    for m in _PRICE_RE.finditer(text):
        raw = m.group(1).replace(",", "").replace("，", "")
        try:
            results.append(float(raw))
        except ValueError:
            pass
    return results


def _parse_ocr_text_to_pricing(ocr_text):
    """
    将 OCR 识别的多行文本解析为报价行列表。
    解析策略：
      1. 逐行扫描，寻找含板数档关键词的「表头行」，建立 tier 顺序；
      2. 对含有 FC 仓号的行，按表头顺序提取金额，映射到对应板数档；
      3. 若无法找到表头，则对含仓号行尝试按固定顺序 [1-4板, 5-14板, 9-13板, 15-28板, 散板] 推断。
    返回: list of {destination_code, pallet_tier, base_price, per_pallet_price}
    """
    lines = [l.strip() for l in ocr_text.splitlines() if l.strip()]
    if not lines:
        return []

    # --- 第一步：寻找表头行，建立 tier 顺序 ---
    tier_order = []
    header_line_idx = -1
    for i, line in enumerate(lines):
        tiers_in_line = []
        for pat, name in _TIER_PATTERNS:
            if pat.search(line):
                tiers_in_line.append((line.index(pat.search(line).group()), name))
        if len(tiers_in_line) >= 2:
            # 按出现位置排序
            tiers_in_line.sort(key=lambda x: x[0])
            tier_order = [t[1] for t in tiers_in_line]
            header_line_idx = i
            break

    # 若没找到表头，用默认顺序
    if not tier_order:
        tier_order = ["1-4板", "5-14板", "9-13板", "15-28板", "散板"]

    # --- 第二步：解析数据行 ---
    results = []
    data_start = header_line_idx + 1 if header_line_idx >= 0 else 0

    for line in lines[data_start:]:
        # 跳过明显是表头或说明行的行
        tier_count = sum(1 for pat, _ in _TIER_PATTERNS if pat.search(line))
        if tier_count >= 2:
            continue  # 另一个表头行，跳过

        fc_codes = _FC_CODE_RE.findall(line)
        if not fc_codes:
            continue

        prices = _extract_prices(line)
        if not prices:
            continue

        # 每个 FC 代码对应同一行的价格
        for fc in fc_codes[:3]:
            fc = fc.upper()
            for j, price in enumerate(prices):
                if j >= len(tier_order):
                    break
                tier = tier_order[j]
                # 判断是全包价还是每板单价：通常 >500 为全包价，否则为每板单价
                if price >= 100:
                    results.append({
                        "destination_code": fc[:10],
                        "pallet_tier": tier,
                        "base_price": price,
                        "per_pallet_price": None,
                    })
                else:
                    results.append({
                        "destination_code": fc[:10],
                        "pallet_tier": tier,
                        "base_price": None,
                        "per_pallet_price": price,
                    })

    return results


def parse_quote_sheet_from_images(ws, sheet_name=""):
    """
    对 worksheet 中的嵌入图片执行 OCR，返回报价行列表。
    格式与 parse_quote_sheet 返回值相同。
    """
    if not hasattr(ws, "_images") or not ws._images:
        return []

    all_rows = []
    img_count = len(ws._images)
    print(f"  [OCR] sheet '{sheet_name}': 发现 {img_count} 张嵌入图片，开始 OCR 识别...")

    for idx, img_obj in enumerate(ws._images):
        try:
            img_bytes = img_obj._data()
        except Exception as e:
            print(f"  [WARN] sheet '{sheet_name}' 图片 {idx+1} 提取失败: {e}")
            continue

        ocr_text = _ocr_image_bytes(img_bytes)
        if not ocr_text or not ocr_text.strip():
            print(f"  [WARN] sheet '{sheet_name}' 图片 {idx+1} OCR 结果为空，跳过")
            continue

        # 打印 OCR 文本摘要（前 120 字符），便于排查
        preview = ocr_text.replace("\n", " ").strip()[:120]
        print(f"  [OCR] 图片 {idx+1} 文本摘要: {preview}")

        rows = _parse_ocr_text_to_pricing(ocr_text)
        if rows:
            print(f"  [OCR] 图片 {idx+1} 解析出 {len(rows)} 条报价记录")
            all_rows.extend(rows)
        else:
            print(f"  [WARN] sheet '{sheet_name}' 图片 {idx+1} 无法解析出结构化报价，已跳过")

    return all_rows


def import_quote(conn, filepath):
    if not os.path.isfile(filepath):
        print("文件不存在:", filepath)
        return 0

    # 启动时检查 OCR 可用性（仅打印一次）
    ocr_ok, ocr_msg = _check_ocr_available()
    if ocr_ok:
        print(f"  [OCR] OCR 引擎就绪: {ocr_msg}")
    else:
        print(f"  [OCR] OCR 不可用，图片 sheet 将被跳过。原因:\n    {ocr_msg}")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_customer = {}
    sheet_images = {}
    all_pricing = []
    ocr_sheet_count = 0

    for sn in wb.sheetnames:
        if not sn or sn in ("无忧达卡派", "Michael", "LZ", "Bayou"):
            continue
        ws = wb[sn]
        cid = "C-Q-" + re.sub(r"[^\w\u4e00-\u9fff]", "", sn)[:10] or "Q"
        sheet_customer[cid] = sn

        # 保留图片 base64（存到 customer details，供后续查看）
        base64_images = []
        if hasattr(ws, "_images"):
            for img in ws._images:
                try:
                    img_data = img._data()
                    fmt = getattr(img, "format", "png")
                    b64 = base64.b64encode(img_data).decode("utf-8")
                    base64_images.append(f"data:image/{fmt};base64,{b64}")
                except Exception:
                    pass
        sheet_images[cid] = base64_images

        # --- 第一步：文本模式解析 ---
        try:
            rows = parse_quote_sheet(ws)
        except Exception as e:
            print(f"  [WARN] sheet '{sn}' 文本解析异常: {e}")
            rows = []

        mode = "文本"

        # --- 第二步：文本为空时，尝试 OCR 图片模式 ---
        if not rows and ocr_ok:
            try:
                rows = parse_quote_sheet_from_images(ws, sheet_name=sn)
                if rows:
                    mode = "图片OCR"
                    ocr_sheet_count += 1
                else:
                    print(f"  [INFO] sheet '{sn}': 文本和图片模式均未解析出报价，已跳过")
            except Exception as e:
                print(f"  [WARN] sheet '{sn}' OCR 模式异常: {e}")
                rows = []
        elif not rows:
            # OCR 不可用且文本为空
            if base64_images:
                print(f"  [INFO] sheet '{sn}': 含 {len(base64_images)} 张图片但 OCR 不可用，已跳过")
            continue

        if not rows:
            continue

        print(f"  [INFO] sheet '{sn}' [{mode}] 解析出 {len(rows)} 条报价记录")
        for r in rows:
            r["customer_id"] = cid
            all_pricing.append(r)

    wb.close()

    for cid, cname in sheet_customer.items():
        details = {"images": sheet_images.get(cid, [])}
        ensure_customer(conn, cid, cname, details=details)

    inserted = 0
    skipped = 0
    with conn.cursor() as cur:
        for p in all_pricing:
            pid = "PM-" + str(uuid.uuid4())[:8].upper()
            try:
                cur.execute("""
                    INSERT INTO pricing_matrices (id, customer_id, destination_code, vehicle_type, pallet_tier, base_price, per_pallet_price, status)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, 'ACTIVE')
                    ON CONFLICT ON CONSTRAINT uq_pricing_matrix DO NOTHING
                """, (
                    pid, p["customer_id"], p["destination_code"], "STRAIGHT_26",
                    p["pallet_tier"], p.get("base_price"), p.get("per_pallet_price")
                ))
                if cur.rowcount:
                    inserted += 1
                else:
                    skipped += 1
            except Exception as e:
                print(f"  [WARN] 报价写入失败 {p.get('destination_code')} / {p.get('pallet_tier')}: {e}")
    conn.commit()

    if ocr_sheet_count:
        print(f"  [OCR] 共有 {ocr_sheet_count} 个 sheet 通过图片 OCR 模式处理")
    if skipped:
        print(f"  [INFO] {skipped} 条已存在的报价记录被跳过（ON CONFLICT DO NOTHING）")
    return inserted


def main():
    url = _load_database_url()
    if not url:
        print("错误: 未找到 DATABASE_URL。")
        print("  方式一: 在 apps/backend/.env 中配置 DATABASE_URL=postgresql://...")
        print("  方式二: 运行前设置环境变量 DATABASE_URL='postgresql://user:pass@host/db?sslmode=require'")
        return 1
    if "..." in url:
        print("错误: DATABASE_URL 中含有占位符 '...'，请填写真实连接串（不要复制文档示例）。")
        print("  请在 apps/backend/.env 中写入完整的 DATABASE_URL，或命令行传入完整 URL。")
        return 1
    print("连接生产库...")
    conn = psycopg2.connect(url)
    try:
        print("导入 2026年转运汇总表.xlsx ...")
        c, i = import_transfer(conn, FILE_TRANSFER)
        print("  转运: 新增 containers", c, ", container_items", i)
        print("导入 合作单位报价卡.xlsx ...")
        p = import_quote(conn, FILE_QUOTE)
        print("  报价: 新增 pricing_matrices", p)
    finally:
        conn.close()
    print("导入完成。")
    return 0


if __name__ == "__main__":
    exit(main())
