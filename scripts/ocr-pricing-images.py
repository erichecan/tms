#!/usr/bin/env python3
"""
Phase 2.2: OCR 合作单位报价卡.xlsx 图片 → partner_pricing_rules
1. 提取嵌入图片
2. 合并 sheet 文本数据
3. 调用 Claude Vision API 识别
4. 写入 partner_pricing_rules

用法:
  python scripts/ocr-pricing-images.py --extract          # 只提取图片到 scripts/pricing-images/
  python scripts/ocr-pricing-images.py --ocr --preview    # OCR + 预览
  python scripts/ocr-pricing-images.py --ocr --commit     # OCR + 写入数据库
  python scripts/ocr-pricing-images.py --sheet "BTL" --ocr --preview  # 单个 sheet
"""

import os
import re
import sys
import json
import base64
import argparse
from io import BytesIO

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ENV_PATH = os.path.join(ROOT, "apps", "backend", ".env")
FILE_QUOTE = os.path.join(ROOT, "合作单位报价卡.xlsx")
IMG_DIR = os.path.join(ROOT, "scripts", "pricing-images")
OCR_CACHE_DIR = os.path.join(ROOT, "scripts", "pricing-ocr-cache")


def _load_env_var(key):
    """Load from env or apps/backend/.env"""
    val = os.environ.get(key, "").strip()
    if val:
        return val
    if os.path.isfile(ENV_PATH):
        with open(ENV_PATH, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line.startswith(f"{key}="):
                    v = line.split("=", 1)[1].strip().strip('"').strip("'")
                    return v
    return ""


# ─── Partner name mapping ─────────────────────────────────────────────────────
PARTNER_MAP = {
    '无忧达卡派': ('无忧达卡派', 'WYDK'),
    'LZ': ('LZ', 'LZ'),
    'ONE': ('ONE', 'ONE'),
    'Michael': ('Michael', 'MCL'),
    'Bayou': ('Bayou', 'BAY'),
    'ESC-XIY': ('ESC-XIY', 'ESCX'),
    'Skyworth': ('Skyworth', 'SKW'),
    'BTL': ('BTL', 'BTL'),
    'Sunny Charm': ('Sunny Charm', 'SC'),
    'Label X': ('Label X', 'LBX'),
    'TEWM': ('TEWM', 'TEWM'),
    'Igloo+DKA': ('Igloo+DKA', 'IGDK'),
    'One Express': ('One Express', 'OEX'),
    'CTCGVTTomXLY': ('CTC/GVT/Tom/XLY', 'CTCG'),
    '祁杰': ('祁杰', 'QJ'),
    '代发客户统一报价表': ('代发统一报价', 'DFTY'),
    'Straightship阿东': ('Straightship阿东', 'SSA'),
    '（缘海）YHWL': ('缘海YHWL', 'YHWL'),
    'NoahHammert': ('NoahHammert', 'NH'),
    'Terry-Ottawa warehouse': ('Terry Ottawa', 'TOW'),
    'Jeff  Don': ('Jeff Don', 'JFD'),
}


def extract_images(wb):
    """Extract all embedded images from image sheets. Returns {sheet_name: [img_paths]}."""
    os.makedirs(IMG_DIR, exist_ok=True)
    result = {}
    for sn, (pname, _) in PARTNER_MAP.items():
        ws = wb[sn]
        imgs = ws._images if hasattr(ws, '_images') else []
        if not imgs:
            result[sn] = []
            continue
        paths = []
        for idx, img in enumerate(imgs):
            try:
                img_data = img._data()
                fmt = getattr(img, 'format', 'png') or 'png'
                safe_name = re.sub(r'[^\w\-]', '_', sn)
                fname = f"{safe_name}_{idx}.{fmt}"
                fpath = os.path.join(IMG_DIR, fname)
                with open(fpath, 'wb') as f:
                    f.write(img_data)
                paths.append(fpath)
            except Exception as e:
                print(f"  [WARN] {sn} img {idx}: {e}")
        result[sn] = paths
    return result


def get_sheet_text(ws):
    """Get all non-empty text from a sheet."""
    lines = []
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        vals = [str(c).strip() for c in row if c is not None and str(c).strip()]
        if vals:
            lines.append(" | ".join(vals))
    return "\n".join(lines)


def ocr_sheet(sn, img_paths, text_context):
    """Call Claude Vision API to OCR a sheet's images + text into structured pricing rules."""
    import anthropic

    api_key = _load_env_var("ANTHROPIC_API_KEY")
    if not api_key:
        print("错误: 未找到 ANTHROPIC_API_KEY（在环境变量或 apps/backend/.env 中设置）")
        sys.exit(1)

    # Check cache first
    os.makedirs(OCR_CACHE_DIR, exist_ok=True)
    safe = re.sub(r'[^\w-]', '_', sn)
    cache_file = os.path.join(OCR_CACHE_DIR, f"{safe}.json")
    if os.path.isfile(cache_file):
        with open(cache_file, 'r', encoding='utf-8') as f:
            cached = json.load(f)
        print(f"  [CACHE] {sn}: 使用缓存 OCR 结果")
        return cached

    client = anthropic.Anthropic(api_key=api_key)

    # Build message content
    content = []
    content.append({
        "type": "text",
        "text": f"这是物流公司的合作单位报价卡中「{sn}」sheet 的数据。请从中提取结构化定价规则。\n\n"
                f"Sheet 文本内容:\n{text_context}\n\n"
                f"下面是嵌入的图片（共 {len(img_paths)} 张）:" if img_paths else
                f"这是物流公司的合作单位报价卡中「{sn}」sheet 的数据。请从中提取结构化定价规则。\n\n"
                f"Sheet 文本内容:\n{text_context}"
    })

    for img_path in img_paths:
        with open(img_path, 'rb') as f:
            img_data = f.read()
        fmt = os.path.splitext(img_path)[1].lstrip('.').lower()
        if fmt == 'jpg':
            fmt = 'jpeg'
        if fmt not in ('png', 'jpeg', 'gif', 'webp'):
            fmt = 'png'
        b64 = base64.b64encode(img_data).decode('utf-8')
        content.append({
            "type": "image",
            "source": {"type": "base64", "media_type": f"image/{fmt}", "data": b64}
        })

    content.append({
        "type": "text",
        "text": """请输出 JSON 数组，每个元素是一条定价规则:
[{
  "destination_warehouse": "仓号/目的地 (如 YYZ3, YOW1, 或城市名)",
  "pricing_type": "matrix/flat/zone/service_fee/contract",
  "transport_mode": "直卡/拖车/小车/大车/53尺/26尺 或 null",
  "pallet_tier_min": 最小板数或null,
  "pallet_tier_max": 最大板数或null,
  "unit_type": "per_trip/per_pallet/per_cbm/per_kg",
  "base_price": 基础价格(数字),
  "unit_price": 单位加价(数字, 如 每板加$10 则填10),
  "surcharges": {"等待费": 50, "远程": 100, ...} 或 {},
  "notes": "附加说明"
}]

要求:
1. 只输出 JSON，不要其他文字
2. 价格是加拿大元 (CAD)
3. 如果图片中有多条线路/目的地的报价，分别列出
4. "首板85 加板35" → base_price=85, unit_price=35, unit_type="per_pallet"
5. 如果是包车/整车价格 → unit_type="per_trip"
6. 如有等待费、远程费等附加费放入 surcharges
7. 仓号格式统一: YYZ3, YOW1, YHM1, YXU1, YGK1 等"""
    })

    print(f"  [OCR] 调用 Claude Vision API for {sn} ({len(img_paths)} imgs)...")
    try:
        response = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=8192,
            messages=[{"role": "user", "content": content}]
        )
        text = response.content[0].text
        # Strip markdown code fences if present
        text_clean = re.sub(r'```json\s*', '', text)
        text_clean = re.sub(r'```\s*$', '', text_clean.strip())
        # Extract JSON from response
        rules = []
        json_match = re.search(r'\[[\s\S]*\]', text_clean)
        if json_match:
            try:
                rules = json.loads(json_match.group())
            except json.JSONDecodeError:
                # Try to fix truncated JSON by closing open brackets
                raw = json_match.group().rstrip().rstrip(',')
                depth_obj = raw.count('{') - raw.count('}')
                depth_arr = raw.count('[') - raw.count(']')
                raw += '}' * depth_obj + ']' * depth_arr
                try:
                    rules = json.loads(raw)
                    print(f"  [WARN] {sn}: Fixed truncated JSON ({len(rules)} rules recovered)")
                except json.JSONDecodeError:
                    print(f"  [WARN] {sn}: JSON parse failed even after fix attempt")
        if not rules:
            # No closing ] — response may be truncated. Try to recover.
            arr_start = text_clean.find('[')
            if arr_start >= 0:
                raw = text_clean[arr_start:].rstrip().rstrip(',')
                last_complete = raw.rfind('},')
                if last_complete > 0:
                    raw = raw[:last_complete + 1] + ']'
                    try:
                        rules = json.loads(raw)
                        print(f"  [RECOVER] {sn}: Recovered {len(rules)} rules from truncated response")
                    except json.JSONDecodeError:
                        pass
        if not rules:
            print(f"  [WARN] {sn}: No rules extracted")
            print(f"  Response: {text[:500]}")
    except Exception as e:
        print(f"  [ERROR] {sn}: OCR failed: {e}")
        rules = []

    # Cache the result
    with open(cache_file, 'w', encoding='utf-8') as f:
        json.dump(rules, f, ensure_ascii=False, indent=2)

    return rules


def ensure_partner(conn, name, short_code):
    """Ensure partner exists, return partner_id."""
    import psycopg2
    with conn.cursor() as cur:
        cur.execute("SELECT id FROM partners WHERE name = %s", (name,))
        row = cur.fetchone()
        if row:
            return row[0]
        cur.execute(
            "INSERT INTO partners (name, short_code, type, status) VALUES (%s, %s, 'carrier', 'ACTIVE') RETURNING id",
            (name, short_code)
        )
        pid = cur.fetchone()[0]
        conn.commit()
        return pid


def insert_ocr_rules(conn, partner_id, rules, sheet_name):
    """Insert OCR-derived rules into partner_pricing_rules."""
    import psycopg2
    count = 0
    with conn.cursor() as cur:
        for r in rules:
            try:
                surcharges = r.get("surcharges", {})
                if isinstance(surcharges, str):
                    surcharges = json.loads(surcharges) if surcharges else {}
                cur.execute("""
                    INSERT INTO partner_pricing_rules
                      (partner_id, pricing_type, destination_warehouse, transport_mode,
                       pallet_tier_min, pallet_tier_max, unit_type,
                       base_price, unit_price, surcharges,
                       source_sheet, source_type, ocr_confidence, notes)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    partner_id,
                    r.get("pricing_type", "matrix"),
                    r.get("destination_warehouse", "ALL"),
                    r.get("transport_mode"),
                    r.get("pallet_tier_min"),
                    r.get("pallet_tier_max"),
                    r.get("unit_type", "per_trip"),
                    r.get("base_price", 0),
                    r.get("unit_price", 0),
                    json.dumps(surcharges),
                    sheet_name,
                    "ocr",
                    0.85,  # default confidence for Claude Vision
                    r.get("notes"),
                ))
                count += 1
            except Exception as e:
                print(f"  [WARN] insert error: {e}")
    conn.commit()
    return count


def main():
    parser = argparse.ArgumentParser(description="OCR 合作单位报价卡图片")
    parser.add_argument("--extract", action="store_true", help="只提取图片")
    parser.add_argument("--ocr", action="store_true", help="执行 OCR")
    parser.add_argument("--preview", action="store_true", help="预览模式")
    parser.add_argument("--commit", action="store_true", help="写入数据库")
    parser.add_argument("--sheet", type=str, help="只处理指定 sheet")
    parser.add_argument("--clear-cache", action="store_true", help="清除 OCR 缓存")
    args = parser.parse_args()

    if not args.extract and not args.ocr:
        print("请指定 --extract 或 --ocr")
        return 1

    if args.clear_cache and os.path.isdir(OCR_CACHE_DIR):
        import shutil
        shutil.rmtree(OCR_CACHE_DIR)
        print("OCR 缓存已清除")

    print(f"读取 {FILE_QUOTE} ...")
    wb = openpyxl.load_workbook(FILE_QUOTE)

    # Step 1: Extract images
    print("提取嵌入图片...")
    img_map = extract_images(wb)
    total_imgs = sum(len(v) for v in img_map.values())
    print(f"提取完成: {total_imgs} 张图片 → {IMG_DIR}/")

    if args.extract and not args.ocr:
        wb.close()
        return 0

    # Step 2: OCR each sheet
    conn = None
    if args.commit:
        import psycopg2
        url = _load_env_var("DATABASE_URL")
        if not url:
            print("错误: 未找到 DATABASE_URL")
            return 1
        conn = psycopg2.connect(url)

    # Reload with data_only for text
    wb.close()
    wb = openpyxl.load_workbook(FILE_QUOTE, data_only=True)

    total_rules = 0
    summary = []

    for sn, (pname, pcode) in PARTNER_MAP.items():
        if args.sheet and sn != args.sheet:
            continue

        ws = wb[sn]
        text_ctx = get_sheet_text(ws)
        img_paths = img_map.get(sn, [])

        if not text_ctx and not img_paths:
            summary.append(f"  EMPTY {sn}: 无文本无图片")
            continue

        # OCR
        rules = ocr_sheet(sn, img_paths, text_ctx)

        if not rules:
            summary.append(f"  EMPTY {sn}: OCR 返回 0 条规则")
            continue

        status = f"  OK    {sn} → {pname} [{pcode}]: {len(rules)} rules"
        if args.preview:
            status += "\n"
            for r in rules[:5]:
                dest = str(r.get('destination_warehouse') or '?')
                bp = r.get('base_price', 0) or 0
                up = r.get('unit_price', 0) or 0
                ut = str(r.get('unit_type') or '?')
                notes = str(r.get('notes') or '')
                line = f"          {dest:12s} base=${bp} unit=${up} ({ut})"
                if notes:
                    line += f" [{str(notes)[:40]}]"
                status += line + "\n"
            if len(rules) > 5:
                status += f"          ... and {len(rules) - 5} more\n"

        summary.append(status)
        total_rules += len(rules)

        if args.commit and conn:
            partner_id = ensure_partner(conn, pname, pcode)
            inserted = insert_ocr_rules(conn, partner_id, rules, sn)
            summary[-1] += f" → DB: {inserted} inserted"

    wb.close()

    print("\n" + "=" * 70)
    print("OCR 报价卡导入摘要")
    print("=" * 70)
    for line in summary:
        print(line)
    print("-" * 70)
    print(f"合计: {total_rules} 条 OCR 规则")
    if args.preview:
        print("\n确认无误后执行: python scripts/ocr-pricing-images.py --ocr --commit")

    if conn:
        conn.close()
    return 0


if __name__ == "__main__":
    exit(main())
